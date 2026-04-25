"""
数据库导出为 Excel
"""
import sqlite3
import os
import sys
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import DB_PATH, OUTPUT_DIR

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def style_header(ws, header_row=1):
    """表头样式"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border


def auto_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        try:
            column_letter = column[0].column_letter
        except AttributeError:
            continue
        for cell in column:
            try:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def export_articles(ws):
    """导出文献主表"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT id, pmid, pmcid, doi, title, abstract, journal, year, pubdate,
                   authors, keywords, mesh_terms, language, article_type,
                   has_fulltext, search_strategy
            FROM articles ORDER BY year DESC, id DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "PMID", "PMCID", "DOI", "标题", "摘要", "期刊", "年份", "发表日期",
               "作者", "关键词", "MeSH词", "语言", "文章类型", "有全文", "检索策略"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_genes(ws):
    """导出基因表（含文献计数）"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT g.id, g.gene_symbol, g.gene_name, g.entrez_id, COUNT(ag.article_id) as article_count
           FROM genes g LEFT JOIN article_gene ag ON g.id = ag.gene_id
           GROUP BY g.id ORDER BY article_count DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "基因符号", "基因名称", "Entrez ID", "关联文献数"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_acupoints(ws):
    """导出穴位表（含文献计数）"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT a.id, a.name_en, a.name_cn, a.code, a.meridian, COUNT(aa.article_id) as article_count
           FROM acupoints a LEFT JOIN article_acupoint aa ON a.id = aa.acupoint_id
           GROUP BY a.id ORDER BY article_count DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "英文名", "中文名", "代码", "经络", "关联文献数"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_diseases(ws):
    """导出疾病表（含文献计数）"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT d.id, d.name, d.mesh_term, d.category, COUNT(ad.article_id) as article_count
           FROM diseases d LEFT JOIN article_disease ad ON d.id = ad.disease_id
           GROUP BY d.id ORDER BY article_count DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "疾病名称", "MeSH词", "分类", "关联文献数"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_article_gene(ws):
    """文献-基因关联表"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT ag.id, a.pmid, a.title, g.gene_symbol
           FROM article_gene ag
           JOIN articles a ON ag.article_id = a.id
           JOIN genes g ON ag.gene_id = g.id
           ORDER BY g.gene_symbol, a.year DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "PMID", "文献标题", "基因符号"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_article_acupoint(ws):
    """文献-穴位关联表"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT aa.id, a.pmid, a.title, apt.name_en, apt.name_cn, apt.code
           FROM article_acupoint aa
           JOIN articles a ON aa.article_id = a.id
           JOIN acupoints apt ON aa.acupoint_id = apt.id
           ORDER BY apt.name_en, a.year DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "PMID", "文献标题", "穴位英文名", "穴位中文名", "代码"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_article_disease(ws):
    """文献-疾病关联表"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT ad.id, a.pmid, a.title, d.name
           FROM article_disease ad
           JOIN articles a ON ad.article_id = a.id
           JOIN diseases d ON ad.disease_id = d.id
           ORDER BY d.name, a.year DESC"""
    ).fetchall()
    conn.close()

    headers = ["ID", "PMID", "文献标题", "疾病"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def export_summary(ws):
    """汇总统计 sheet"""
    conn = get_conn()

    # 总体统计
    stats = {
        "收录文献总数": conn.execute("SELECT COUNT(*) FROM articles").fetchone()[0],
        "关联基因总数": conn.execute("SELECT COUNT(*) FROM genes").fetchone()[0],
        "关联穴位总数": conn.execute("SELECT COUNT(*) FROM acupoints").fetchone()[0],
        "关联疾病总数": conn.execute("SELECT COUNT(*) FROM diseases").fetchone()[0],
        "基因-文献关联": conn.execute("SELECT COUNT(*) FROM article_gene").fetchone()[0],
        "穴位-文献关联": conn.execute("SELECT COUNT(*) FROM article_acupoint").fetchone()[0],
        "疾病-文献关联": conn.execute("SELECT COUNT(*) FROM article_disease").fetchone()[0],
        "有全文文献": conn.execute("SELECT COUNT(*) FROM articles WHERE has_fulltext=1").fetchone()[0],
    }
    conn.close()

    ws.append(["针灸穴位-基因关联数据库 汇总统计"])
    ws.append([])
    ws.append(["指标", "数值"])
    for k, v in stats.items():
        ws.append([k, v])

    ws.append([])
    ws.append(["Top 20 高频基因"])
    ws.append(["排名", "基因", "关联文献数"])

    conn = get_conn()
    top_genes = conn.execute(
        """SELECT g.gene_symbol, COUNT(*) as cnt FROM genes g
           JOIN article_gene ag ON g.id = ag.gene_id
           GROUP BY g.id ORDER BY cnt DESC LIMIT 20"""
    ).fetchall()
    for i, (gene, cnt) in enumerate(top_genes, 1):
        ws.append([i, gene, cnt])

    ws.append([])
    ws.append(["Top 20 高频穴位"])
    ws.append(["排名", "穴位", "关联文献数"])

    top_acupoints = conn.execute(
        """SELECT a.name_en, COUNT(*) as cnt FROM acupoints a
           JOIN article_acupoint aa ON a.id = aa.acupoint_id
           GROUP BY a.id ORDER BY cnt DESC LIMIT 20"""
    ).fetchall()
    for i, (apt, cnt) in enumerate(top_acupoints, 1):
        ws.append([i, apt, cnt])

    ws.append([])
    ws.append(["Top 15 高频疾病"])
    ws.append(["排名", "疾病", "关联文献数"])

    top_diseases = conn.execute(
        """SELECT d.name, COUNT(*) as cnt FROM diseases d
           JOIN article_disease ad ON d.id = ad.disease_id
           GROUP BY d.id ORDER BY cnt DESC LIMIT 15"""
    ).fetchall()
    for i, (disease, cnt) in enumerate(top_diseases, 1):
        ws.append([i, disease, cnt])

    ws.append([])
    ws.append(["年份分布"])
    ws.append(["年份", "文献数"])
    years = conn.execute(
        "SELECT year, COUNT(*) FROM articles WHERE year IS NOT NULL GROUP BY year ORDER BY year"
    ).fetchall()
    for y, cnt in years:
        ws.append([y, cnt])
    conn.close()

    # 样式
    title_font = Font(size=16, bold=True, color="4472C4")
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center")

    for row in [3, 6, 27, 49]:
        for cell in ws[row]:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

    auto_width(ws)


def export_network(ws):
    """穴位-基因共现网络表"""
    conn = get_conn()
    rows = conn.execute(
        """SELECT a.name_en as acupoint, g.gene_symbol as gene, COUNT(*) as cooccur_count
           FROM article_acupoint aa
           JOIN article_gene ag ON aa.article_id = ag.article_id
           JOIN acupoints a ON aa.acupoint_id = a.id
           JOIN genes g ON ag.gene_id = g.id
           GROUP BY a.id, g.id HAVING cooccur_count >= 2
           ORDER BY cooccur_count DESC LIMIT 200"""
    ).fetchall()
    conn.close()

    headers = ["穴位", "基因", "共现文献数"]
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    style_header(ws)
    auto_width(ws)
    ws.freeze_panes = "A2"


def main():
    print("[EXCEL] 开始导出数据库到 Excel...")
    wb = openpyxl.Workbook()

    # 删除默认 sheet
    wb.remove(wb.active)

    # 创建各 sheet
    sheets = [
        ("汇总统计", export_summary),
        ("文献主表", export_articles),
        ("基因统计", export_genes),
        ("穴位统计", export_acupoints),
        ("疾病统计", export_diseases),
        ("文献-基因关联", export_article_gene),
        ("文献-穴位关联", export_article_acupoint),
        ("文献-疾病关联", export_article_disease),
        ("共现网络", export_network),
    ]

    for sheet_name, export_func in sheets:
        ws = wb.create_sheet(title=sheet_name)
        export_func(ws)
        print(f"  [OK] {sheet_name}: {ws.max_row-1} 行")

    output_path = os.path.join(OUTPUT_DIR, "acupoint_gene_database.xlsx")
    wb.save(output_path)
    print(f"\n[EXCEL] 导出完成: {output_path}")
    print(f"[EXCEL] 文件大小: {os.path.getsize(output_path) / 1024:.1f} KB")


if __name__ == "__main__":
    main()
